"""
Export Service
==============

Servicio de exportación a diferentes formatos.
"""

import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)


class ExportService:
    """Servicio de exportación."""
    
    def __init__(self, export_dir: str = "exports"):
        """
        Inicializar servicio de exportación.
        
        Args:
            export_dir: Directorio para exports
        """
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        self._logger = logger
    
    def export_to_pdf(
        self,
        artist_id: str,
        data: Dict[str, Any],
        title: str = "Reporte"
    ) -> str:
        """
        Exportar a PDF.
        
        Args:
            artist_id: ID del artista
            data: Datos a exportar
            title: Título del documento
        
        Returns:
            Ruta del archivo PDF
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{artist_id}_{timestamp}.pdf"
            filepath = self.export_dir / filename
            
            doc = SimpleDocTemplate(str(filepath), pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            story.append(Paragraph(title, styles['Title']))
            story.append(Spacer(1, 0.2*inch))
            
            # Contenido básico
            story.append(Paragraph(f"Artista ID: {artist_id}", styles['Normal']))
            story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            # Agregar datos como tabla si es posible
            if isinstance(data, dict):
                table_data = [["Campo", "Valor"]]
                for key, value in data.items():
                    table_data.append([str(key), str(value)])
                
                table = Table(table_data)
                story.append(table)
            
            doc.build(story)
            self._logger.info(f"Exported to PDF: {filepath}")
            return str(filepath)
        
        except ImportError:
            self._logger.warning("reportlab not installed, PDF export disabled")
            raise ImportError("reportlab library required for PDF export")
        except Exception as e:
            self._logger.error(f"Error exporting to PDF: {str(e)}")
            raise
    
    def export_to_excel(
        self,
        artist_id: str,
        data: Dict[str, Any],
        sheet_name: str = "Datos"
    ) -> str:
        """
        Exportar a Excel.
        
        Args:
            artist_id: ID del artista
            data: Datos a exportar
            sheet_name: Nombre de la hoja
        
        Returns:
            Ruta del archivo Excel
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{artist_id}_{timestamp}.xlsx"
            filepath = self.export_dir / filename
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Encabezados
            row = 1
            ws.cell(row=row, column=1, value="Campo")
            ws.cell(row=row, column=2, value="Valor")
            
            # Datos
            row = 2
            for key, value in data.items():
                ws.cell(row=row, column=1, value=str(key))
                ws.cell(row=row, column=2, value=str(value))
                row += 1
            
            wb.save(str(filepath))
            self._logger.info(f"Exported to Excel: {filepath}")
            return str(filepath)
        
        except ImportError:
            self._logger.warning("openpyxl not installed, Excel export disabled")
            raise ImportError("openpyxl library required for Excel export")
        except Exception as e:
            self._logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def export_events_to_ical(
        self,
        artist_id: str,
        events: List[Dict[str, Any]]
    ) -> str:
        """
        Exportar eventos a formato iCal.
        
        Args:
            artist_id: ID del artista
            events: Lista de eventos
        
        Returns:
            Ruta del archivo iCal
        """
        try:
            from icalendar import Calendar, Event as ICalEvent
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"events_{artist_id}_{timestamp}.ics"
            filepath = self.export_dir / filename
            
            cal = Calendar()
            cal.add('prodid', '-//Artist Manager AI//EN')
            cal.add('version', '2.0')
            
            for event_data in events:
                event = ICalEvent()
                event.add('summary', event_data.get('title', ''))
                event.add('description', event_data.get('description', ''))
                event.add('dtstart', datetime.fromisoformat(event_data.get('start_time', '')))
                event.add('dtend', datetime.fromisoformat(event_data.get('end_time', '')))
                if event_data.get('location'):
                    event.add('location', event_data['location'])
                
                cal.add_component(event)
            
            with open(filepath, 'wb') as f:
                f.write(cal.to_ical())
            
            self._logger.info(f"Exported to iCal: {filepath}")
            return str(filepath)
        
        except ImportError:
            self._logger.warning("icalendar not installed, iCal export disabled")
            raise ImportError("icalendar library required for iCal export")
        except Exception as e:
            self._logger.error(f"Error exporting to iCal: {str(e)}")
            raise




