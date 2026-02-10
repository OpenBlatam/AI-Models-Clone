"""Excel Converter - Convert Markdown to Excel format"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from typing import Dict, Any, Optional

from .base_converter import BaseConverter
from ...utils.chart_generator import ChartGenerator


class ExcelConverter(BaseConverter):
    """Convert Markdown to Excel (.xlsx) format"""
    
    def __init__(self):
        self.chart_generator = ChartGenerator()
    
    async def convert(
        self,
        parsed_content: Dict[str, Any],
        output_path: str,
        include_charts: bool = True,
        include_tables: bool = True,
        custom_styling: Optional[Dict[str, Any]] = None
    ) -> None:
        """Convert to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Content"
        
        row = 1
        
        # Add title
        if parsed_content.get("title"):
            ws.merge_cells(f'A{row}:D{row}')
            title_cell = ws[f'A{row}']
            title_cell.value = parsed_content["title"]
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 2
        
        # Add headings
        for heading in parsed_content.get("headings", []):
            col = heading["level"] - 1
            cell = ws.cell(row=row, column=col+1)
            cell.value = heading["text"]
            cell.font = Font(size=14-heading["level"], bold=True)
            row += 1
        
        # Add tables
        if include_tables:
            for table in parsed_content.get("tables", []):
                # Add headers
                for col_idx, header in enumerate(table["headers"], start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center')
                
                row += 1
                
                # Add data rows
                for data_row in table["rows"]:
                    for col_idx, value in enumerate(data_row, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = value
                        cell.alignment = Alignment(horizontal='left', vertical='top')
                    row += 1
                
                # Auto-adjust column widths
                for col_idx in range(1, len(table["headers"]) + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0
                    for cell in ws[col_letter]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                row += 2
        
        # Add charts if tables exist and charts are enabled
        if include_charts and parsed_content.get("tables"):
            chart_row = row
            for idx, table in enumerate(parsed_content.get("tables", [])):
                if len(table["rows"]) > 0:
                    try:
                        # Auto-detect best chart type
                        chart_type = self.chart_generator.auto_detect_chart_type(table)
                        
                        # Create chart based on detected type
                        if chart_type == "bar":
                            chart = BarChart()
                        elif chart_type == "line":
                            chart = LineChart()
                        elif chart_type == "pie":
                            chart = PieChart()
                        else:
                            chart = BarChart()
                        
                        chart.title = f"{table['headers'][0] if table.get('headers') else 'Chart'} {idx + 1}"
                        chart.style = 10
                        
                        # Find numeric data
                        data_start_row = chart_row - len(table["rows"]) - 1
                        data_end_row = chart_row - 1
                        
                        # Try to use first column as categories, second as values
                        if len(table["headers"]) >= 2:
                            categories = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
                            values = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
                            chart.add_data(values, titles_from_data=True)
                            chart.set_categories(categories)
                            chart.width = 15
                            chart.height = 10
                            ws.add_chart(chart, f"F{chart_row}")
                            chart_row += 15
                    except Exception as e:
                        # Skip chart if there's an error
                        pass
        
        # Add paragraphs as text
        for para in parsed_content.get("paragraphs", []):
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = para
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            row += 1
        
        # Save workbook
        wb.save(output_path)

